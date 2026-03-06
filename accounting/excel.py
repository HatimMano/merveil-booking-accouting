"""Create an Excel workbook (.xlsx) with 3 sheets: Journal, Anomalies, Matrice de vérification."""

import logging
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from accounting.entries import AccountingEntry
from models.reservation import Reservation
from validators.anomalies import Anomaly, Severity

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

_BLUE_FILL   = PatternFill("solid", fgColor="2F5496")
_ORANGE_FILL = PatternFill("solid", fgColor="C55A11")   # BLOCKING
_YELLOW_FILL = PatternFill("solid", fgColor="FFD966")   # WARNING
_GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")   # TOTAL row

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WHITE_BOLD  = Font(bold=True, color="FFFFFF")
_BLACK_BOLD  = Font(bold=True)


def _header_row(ws, titles: List[str]) -> None:
    """Write and style a header row on the active row of ws."""
    ws.append(titles)
    for cell in ws[ws.max_row]:
        cell.font = _WHITE_BOLD
        cell.fill = _BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _set_col_widths(ws, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt(value: Optional[Decimal]) -> str:
    """Format a Decimal for display; empty string if None or zero."""
    if value is None:
        return ""
    return f"{value:.2f}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_journal(ws, entries: List[AccountingEntry]) -> None:
    ws.title = "Journal"
    _header_row(ws, ["Journal", "Date", "Réf. pièce", "Compte", "Libellé", "Débit", "Crédit"])
    _set_col_widths(ws, [10, 12, 12, 16, 70, 14, 14])

    for entry in entries:
        ws.append([
            entry.journal,
            entry.date,          # Python date object → Excel date cell
            entry.ref_piece,
            entry.account,
            entry.label,
            float(entry.debit)   if entry.debit  is not None else None,
            float(entry.credit)  if entry.credit is not None else None,
        ])
        row = ws[ws.max_row]
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
        # Date column — force date display format
        row[1].number_format = "DD/MM/YYYY"
        # Amount columns — right-align with 2 decimal places
        for cell in (row[5], row[6]):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if cell.value is not None:
                cell.number_format = "#,##0.00"


def _build_anomalies(ws, anomalies: List[Anomaly]) -> None:
    ws.title = "Anomalies"
    _header_row(ws, ["Sévérité", "Type", "Message", "Fichier source", "Réf. réservation"])
    _set_col_widths(ws, [12, 26, 80, 40, 20])

    _sev_fill = {
        Severity.BLOCKING: _ORANGE_FILL,
        Severity.WARNING:  _YELLOW_FILL,
    }

    for a in anomalies:
        ws.append([
            a.severity,
            a.type,
            a.message,
            a.source_file,
            a.reservation_ref or "",
        ])
        row = ws[ws.max_row]
        fill = _sev_fill.get(a.severity)
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    if not anomalies:
        ws.append(["Aucune anomalie détectée."])
        ws[ws.max_row][0].font = Font(italic=True, color="70AD47")


def _build_matrix(ws, reservations: List[Reservation]) -> None:
    ws.title = "Matrice de vérification"
    _header_row(ws, ["ID Appartement", "Ref Appart", "SUM Débit", "SUM Crédit", "SUM Net"])
    _set_col_widths(ws, [20, 16, 14, 14, 14])

    # Group by (code_comptable, ref_appart)
    groups: dict = {}
    for r in reservations:
        key = (r.code_comptable or "", r.ref_appart)
        if key not in groups:
            groups[key] = {"debit": Decimal("0"), "credit": Decimal("0"), "net": Decimal("0")}
        groups[key]["debit"]  += r.amount
        groups[key]["credit"] += abs(r.commission) + abs(r.payment_charge)
        groups[key]["net"]    += r.net

    total_debit  = Decimal("0")
    total_credit = Decimal("0")
    total_net    = Decimal("0")

    for (code_comptable, ref_appart), t in sorted(groups.items(), key=lambda x: x[0][0]):
        ws.append([
            code_comptable,
            ref_appart,
            _fmt(t["debit"]),
            _fmt(t["credit"]),
            _fmt(t["net"]),
        ])
        row = ws[ws.max_row]
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
        for cell in (row[2], row[3], row[4]):
            cell.alignment = Alignment(horizontal="right", vertical="center")

        total_debit  += t["debit"]
        total_credit += t["credit"]
        total_net    += t["net"]

    # TOTAL row
    ws.append(["TOTAL", "", _fmt(total_debit), _fmt(total_credit), _fmt(total_net)])
    row = ws[ws.max_row]
    for cell in row:
        cell.font   = _BLACK_BOLD
        cell.fill   = _GREY_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")
    row[0].alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_excel_workbook(
    entries: List[AccountingEntry],
    anomalies: List[Anomaly],
    reservations: List[Reservation],
    output_path: Path,
) -> None:
    """
    Create an Excel workbook with 3 sheets:
      1. Journal               — all accounting entries
      2. Anomalies             — validation warnings / errors
      3. Matrice de vérification — per-apartment debit / credit / net summary

    Args:
        entries:       AccountingEntry objects from generate_entries().
        anomalies:     All Anomaly objects collected during the run.
        reservations:  Processed reservations (code_comptable already set).
        output_path:   Destination .xlsx file path.
    """
    wb = openpyxl.Workbook()

    _build_journal(wb.active, entries)
    _build_anomalies(wb.create_sheet(), anomalies)
    _build_matrix(wb.create_sheet(), reservations)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel workbook written to: %s (3 sheets)", output_path)
